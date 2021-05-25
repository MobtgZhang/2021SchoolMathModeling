import csv
import os
import json
import re
from openpyxl import Workbook
from bs4 import BeautifulSoup
from tqdm import tqdm

from .filter import filter_chinese,jieba_segment

def combine(logger,read_file,save_file):
    '''
    合并数据文件,将22个数据文件全部合并起来
    注意原始文件是GBK编码的,现在将其转化为UTF-8编码进行操作
    :param file_path: 读取的数据文件
    :param save_file: 保存的数据文件
    :return: 无
    '''
    filenames = os.listdir(read_file)
    filenames = sorted(filenames,key=lambda x:int(re.findall('NLP_Corpus_(.*).txt',x)[0]))
    with open(save_file,mode="w",encoding="utf-8") as wfp:
        for file in filenames:
            tmp_file = os.path.join(read_file,file)
            logger.info("Read %s"%tmp_file)
            with open(tmp_file, mode="r", encoding="gbk") as rfp:
                for line in rfp:
                    wfp.write(line.strip()+"\n")
    logger.info("Saved file %s"%save_file)
def save_csv_file(all_data,save_file,encoding="utf-8"):
    '''
    将数据保存为csv文件
    :param all_data: 所有的数据集合
    :param save_file: 保存的文件
    :param encoding: 文件的编码
    :return: 无
    '''
    with open(save_file, mode="w", encoding=encoding) as wfp:
        writer = csv.writer(wfp)
        writer.writerow(all_data[0].keys())
        for item in all_data:
            writer.writerow(list(item.values()))
def save_json_file(all_data,save_file,encoding="utf-8"):
    '''
    将数据保存为json文件
    :param all_data: 所有的数据集合
    :param save_file: 保存的文件
    :param encoding: 文件的编码
    :return: 无
    '''
    with open(save_file, mode="w", encoding=encoding) as wfp:
        json.dump(all_data, wfp)
def save_xlsx_file(all_data,save_file):
    '''
    将数据保存为json文件
    :param all_data: 所有的数据集合
    :param save_file: 保存的文件
    :return: 无
    '''
    # 创建一个Workbook对象
    workbook = Workbook()
    # 如果不指定sheet索引和表名，默认在第二张表位置新建表名sheet1
    workbook.create_sheet(index=0, title="原始数据")
    # 获取当前活跃的sheet，默认为第一张sheet
    booksheet = workbook.active
    # 设置worksheet样式
    headers = list(all_data[0].keys())
    for j, word in enumerate(headers):
        booksheet.cell(1, j+1).value = word
    for i, item in enumerate(all_data):
        for j, word in enumerate(item):
            booksheet.cell(i + 2, j+1).value = item[word]
    # 保存为工作簿
    workbook.save(save_file)
def preprocess(logger,args,type_name = 'json'):
    '''
    用bs4工具将文件转化为json,csv,excel等等可读取文件
    :param args: 参数
    :param type: 文件保存的类型
    :return: 无
    '''
    type_name = type_name.lower()
    assert type_name in ['json', 'csv', 'xlsx','all']
    raw_file = os.path.join(args.processed_path, 'NLP_Corpus.txt')
    save_json = os.path.join(args.processed_path, "NLP_Corpus.json")
    save_csv = os.path.join(args.processed_path, "NLP_Corpus.csv")
    save_xlsx = os.path.join(args.processed_path, "NLP_Corpus.xlsx")
    with open(raw_file,mode="r",encoding="utf-8") as rfp:
        all_data = []
        soup = BeautifulSoup(rfp,'lxml')
        for item in tqdm(soup.find_all("record"),desc="bs4 filter processing"):
            idx = item.id.get_text().strip()
            article = item.article.get_text().strip()
            discuss = item.discuss.get_text().strip()
            insertTime = item.inserttime.get_text().strip()
            origin = item.origin.get_text().strip()
            person_id = item.person_id.get_text().strip()
            time = item.time.get_text().strip()
            transmit = item.transmit.get_text().strip()
            tmp_dict = {
                "id": idx,
                "article": article,
                "discuss": discuss,
                "insertTime": insertTime,
                "origin": origin,
                "person_id": person_id,
                "time": time,
                "transmit": transmit
            }
            all_data.append(tmp_dict)
    if type_name == 'json':
        save_json_file(all_data,save_json)
        logger.info("Save in file:%s" % save_json)
    elif type_name == 'csv':
        save_csv_file(all_data,save_csv)
        logger.info("Save in file:%s" % save_csv)
    elif type_name == 'xlsx':
        save_xlsx_file(all_data,save_xlsx)
        logger.info("Save in file:%s" % save_xlsx)
    elif type_name == 'all':
        save_csv_file(all_data, save_csv)
        logger.info("Save in file:%s" % save_csv)
        save_json_file(all_data, save_json)
        logger.info("Save in file:%s" % save_json)
        save_xlsx_file(all_data, save_xlsx)
        logger.info("Save in file:%s" % save_xlsx)
    else:
        raise Exception("Unknown type:%s"%(type_name))
def get_person_filter(save_json,save_person_filter,stop_words_file,jieba = True,singleWords=False):
    all_data = {}
    # 筛选出不重复的文章
    with open(save_json,mode="r",encoding="utf-8") as rfp:
        raw_data = json.load(rfp)
        for item in tqdm(raw_data,'creating person data'):
            person_id = item['person_id']
            article = item['article']
            if article.strip() == '':
                continue
            time1 = item['time']
            time2 = item['insertTime']
            transmit = item['transmit']
            discuss = item['discuss']
            idx = item['id']
            if person_id in all_data:
                if article in all_data[person_id]:
                    continue
                else:
                    all_data[person_id][article] = {
                        "id": idx,
                        "time":time1,
                        "transmit":transmit,
                        "discuss":discuss
                    }
            else:
                all_data[person_id] = {}
                all_data[person_id][article] = {
                    "id":idx,
                    "time": time1,
                    "transmit": transmit,
                    "discuss": discuss
                }
    re_filter_data = {}
    for person_id in tqdm(all_data,'exchange,filter and segment data'):
        re_filter_data[person_id] = {}
        for article in all_data[person_id]:
            idx = all_data[person_id][article]['id']
            time1 = all_data[person_id][article]['time']
            transmit = all_data[person_id][article]['transmit']
            discuss = all_data[person_id][article]['discuss']
            if article.strip() == '':
                continue
            document = filter_chinese(article)
            if document != '':
                if jieba:
                    document = jieba_segment(document, stop_words_file, singleWords)
                    if len(document) == 0:
                        continue
                    re_filter_data[person_id][idx] = {
                        "article": document,
                        "time": time1,
                        "transmit": transmit,
                        "discuss": discuss
                    }
    del all_data
    with open(save_person_filter,mode="w",encoding="utf-8") as wfp:
        json.dump(re_filter_data,wfp)
